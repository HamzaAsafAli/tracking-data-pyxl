import os
import sys

import openpyxl
from openpyxl.styles import Font

try:
    from openpyxl.cell import get_column_letter, column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string


class Solution:
    def __init__(self):
        if os.path.isfile('track_temp.xlsx'):
            self.wb = openpyxl.load_workbook('track_temp.xlsx')
            self.sheet = self.wb['Expenses']
            self.sheet['C6'].value = '=SUM(C2:C5)'
            self.sheet['D6'].value = '=SUM(D2:D5)'
            self.sheet['E6'].value = '=SUM(E2:E5)'
        else:
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.active
            fontObj1 = Font(name='Times New Roman', bold=True)
            # makes the column wider
            self.sheet.column_dimensions['A'].width = 20
            self.sheet.column_dimensions['B'].width = 20
            self.sheet.column_dimensions['C'].width = 20
            self.sheet.column_dimensions['D'].width = 20
            self.sheet.column_dimensions['E'].width = 20
            # set up the basic structure
            self.sheet.title = 'Expenses'
            self.sheet['A1'].font = fontObj1
            self.sheet['A1'] = 'Spend Categories'
            self.sheet['A2'] = 'Transportation'
            self.sheet['A3'] = 'Retail & Groceries'
            self.sheet['A4'] = 'Entertainment'
            self.sheet['A5'] = 'Restaurants'
            self.sheet['A6'].font = fontObj1
            self.sheet['A6'] = 'Total ($)'
            self.sheet['B1'].font = fontObj1
            self.sheet['B1'] = "Transactions"
            self.sheet['C1'].font = fontObj1
            self.sheet['C1'] = "Amount ($)"
            self.sheet['D1'].font = fontObj1
            self.sheet['D1'] = "Budget ($)"
            self.sheet['E1'].font = fontObj1
            self.sheet['E1'] = "Difference ($)"
            self.wb.save('track_temp.xlsx')

    def updater(self):

        print("Enter digit for the spend category: (1) Transportation, (2) Retail & Groceries, (3) Entertainment "
              "and (4) Restaurants ")
        category = input("Enter corresponding spend category digit or type exit to terminate: ")

        if category == "1":
            # working with columns B - E, row 2
            print("You are entering expenses for Transportation")
            check = input("Proceed Y/N? ")
            if check == "Y" or check == "y":
                self.sheet.cell(row=2, column=2).value = int(input("Enter number of transactions: "))
                self.sheet.cell(row=2, column=3).value = int(input("Enter amount: "))
                self.sheet.cell(row=2, column=4).value = int(input("Enter your budget: "))
                self.sheet.cell(row=2, column=5).value = '=D2-C2'
                #self.sheet['C6'].value = '=SUM(C2:C5)'
                self.wb.save('track_temp.xlsx')
            self.updater()
        elif category == "2":
            # working with columns B - E, row 3
            print("You are entering expenses for Retail & Groceries")
            check = input("Proceed Y/N? ")
            if check == "Y" or check == "y":
                self.sheet['B3'] = int(input("Enter number of transactions: "))
                self.sheet.cell(row=3, column=3).value = int(input("Enter amount: "))
                self.sheet.cell(row=3, column=4).value = int(input("Enter your budget: "))
                self.sheet.cell(row=3, column=5).value = '=D3-C3'
                #self.sheet['C6'].value = '=SUM(C2:C5)'
                self.wb.save('track_temp.xlsx')
            self.updater()
        elif category == "3":
            # working with columns B - E, row 4
            print("You are entering expenses for Restaurants")
            check = input("Proceed Y/N? ")
            if check == "Y" or check == "y":
                self.sheet['B4'] = int(input("Enter number of transactions: "))
                self.sheet.cell(row=4, column=3).value = int(input("Enter amount: "))
                self.sheet.cell(row=4, column=4).value = int(input("Enter your budget: "))
                self.sheet.cell(row=4, column=5).value = '=D4-C4'
                #self.sheet['C6'].value = '=SUM(C2:C5)'
                self.wb.save('track_temp.xlsx')
            self.updater()
        elif category == "4":
            # working with columns B - E, row 5
            print("You are entering expenses for Entertainment")
            check = input("Proceed Y/N? ")
            if check == "Y" or check == "y":
                self.sheet['B5'] = int(input("Enter number of transactions: "))
                self.sheet.cell(row=5, column=3).value = int(input("Enter amount: "))
                self.sheet.cell(row=5, column=4).value = int(input("Enter your budget: "))
                self.sheet.cell(row=5, column=5).value = '=D5-C5'
                #self.sheet['C6'].value = '=SUM(C2:C5)'
                self.wb.save('track_temp.xlsx')
            self.updater()
        elif category == "exit" or category == "Exit":
            sys.exit()
        else:
            self.updater()

        self.wb.save('track_temp.xlsx')


a = Solution()
a.updater()



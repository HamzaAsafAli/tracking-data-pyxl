import openpyxl

try:
    from openpyxl.cell import get_column_letter, column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
tuple(sheet['A1':'C3'])

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')
# print(sheet['A1'])
# print(sheet['A1'].value)
# c = sheet['B1']  #cell
# print(c.value)
# print('Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value)
# print('Cell ' + c.coordinate + ' is ' + c.value)
# print(sheet['C1'].value)
# print(c.column)
# print(get_column_letter(c.column))
# print(column_index_from_string('A'))
# print(sheet['C1'].column)
# for i in range(1, 8, 2):
#     print(i, sheet.cell(row=i, column=2).value)
#
# print(sheet.max_column)
#print(type(activesheet))